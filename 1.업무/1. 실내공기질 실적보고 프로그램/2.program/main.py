
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font

st.set_page_config(page_title="실내공기질 데이터 처리 앱", layout="wide")

def to_excel_bytes(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return output

def format_excel(file_bytes):
    wb = load_workbook(file_bytes)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(size=10)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 11
    formatted = BytesIO()
    wb.save(formatted)
    formatted.seek(0)
    return formatted

st.title("📊 실내공기질 데이터 처리 앱 (1~6단계)")

uploaded_note = st.file_uploader("1️⃣ 실험노트 파일 업로드", type="xlsx", key="note")
uploaded_stat = st.file_uploader("2️⃣ 조회통계 파일 업로드", type="xlsx", key="stat")

if uploaded_note and uploaded_stat:
    data_df = pd.read_excel(uploaded_note)
    result_df = pd.read_excel(uploaded_stat)

    # 1단계: 병합
    merged_df = pd.merge(data_df, result_df, on='접수번호', how='inner')
    file1 = to_excel_bytes(merged_df)
    st.download_button("📥 1.병합 데이터.xlsx", file1, file_name="1.병합 데이터.xlsx")
    st.dataframe(merged_df.head())

    # 2단계: 실내공기질 필터링
    filtered_df = merged_df[merged_df['검체유형_x'].str.contains('실내공기질', na=False)].copy()
    file2 = to_excel_bytes(filtered_df)
    st.download_button("📥 2.실내공기질 필터링 데이터.xlsx", file2, file_name="2.실내공기질 필터링 데이터.xlsx")
    st.dataframe(filtered_df.head())

    # 3단계: 변수 추출
    filtered_df['시설군'] = filtered_df['검체유형_x'].str.split('/').str[2]
    selected_columns = [
        '의뢰기관', '시설군', '배출시설', '시설명', '채취장소_x', '시료명_x',
        'PM10', 'PM2.5', 'CO2', '폼알데하이드', '부유세균', '일산화탄소',
        '라돈', '라돈(밀폐)', '벤젠', '톨루엔', '에틸벤젠', '자일렌', '스틸렌',
        '부적합항목', '확인일', '검사결과', '접수번호', '접수일자'
    ]
    extracted_df = filtered_df[selected_columns].copy()
    file3 = to_excel_bytes(extracted_df)
    st.download_button("📥 3.추출 데이터.xlsx", file3, file_name="3.추출 데이터.xlsx")
    st.dataframe(extracted_df.head())

    # 공통 전처리
    df = extracted_df.rename(columns={
        '의뢰기관': '시군명',
        '배출시설': '세부시설군',
        '채취장소_x': '주소',
        'PM10': '미세먼지(PM10)',
        'PM2.5': '초미세먼지(PM2.5)',
        'CO2': '이산화탄소',
        '부유세균': '총부유세균',
        '시료명_x': '시료명'
    })
    df['접수번호_10자리'] = df['접수번호'].astype(str).str[:10]
    df['채취지점'] = df.groupby('시설명')['시료명'].transform(lambda x: ', '.join(x.unique()))
    df_clean = df.copy()

    # 4단계: 기존/신축공동주택 필터링
    apart_df = df_clean[df_clean['시설군'].isin(['기존공동주택', '신축공동주택'])].copy()
    file4 = to_excel_bytes(apart_df)
    st.download_button("📥 4.기존신축공동주택 데이터.xlsx", file4, file_name="4.기존신축공동주택 데이터.xlsx")
    st.dataframe(apart_df.head())

    # 5단계: 다중이용시설 처리
    for col in ['미세먼지(PM10)', '초미세먼지(PM2.5)', '이산화탄소', '폼알데하이드', '총부유세균', '일산화탄소', '라돈']:
        if col in df_clean.columns:
            df_clean[col] = df_clean[col].replace("불검출", None)
            df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')

    avg_cols = ['미세먼지(PM10)', '초미세먼지(PM2.5)', '이산화탄소', '폼알데하이드', '총부유세균', '일산화탄소', '라돈']
    avg_df = df_clean.groupby(['시설명', '접수번호_10자리'])[avg_cols].mean().reset_index()
    reduced = df_clean.drop(columns=avg_cols, errors='ignore').drop_duplicates(['시설명', '접수번호_10자리'])
    merged = avg_df.merge(reduced, on=['시설명', '접수번호_10자리'])
    merged.loc[merged['검사결과'] == '적합', '부적합항목'] = ""
    multi_df = merged[~merged['시설군'].isin(['기존공동주택', '신축공동주택'])].copy()
    file5 = to_excel_bytes(multi_df)
    st.download_button("📥 5.다중이용시설 데이터.xlsx", file5, file_name="5.다중이용시설 데이터.xlsx")
    st.dataframe(multi_df.head())

    # 6단계: 통합
    apart_df['확인일'] = pd.to_datetime(apart_df['확인일'], errors='coerce').dt.date
    multi_df['확인일'] = pd.to_datetime(multi_df['확인일'], errors='coerce').dt.date
    apart_df.sort_values('확인일', inplace=True)
    multi_df.sort_values('확인일', inplace=True)
    multi_df.drop(columns='접수번호', inplace=True, errors='ignore')

    def to_excel_multi(dfs, names):
        out = BytesIO()
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            for df, name in zip(dfs, names):
                df.to_excel(writer, index=False, sheet_name=name)
        out.seek(0)
        return out

    combined = to_excel_multi([multi_df, apart_df], ['다중이용시설', '기존신축공동주택'])
    styled = format_excel(combined)
    st.download_button("📥 6.통합 데이터.xlsx", styled, file_name="6.통합 데이터.xlsx")

else:
    st.info("👆 실험노트와 조회통계 파일을 모두 업로드해주세요.")