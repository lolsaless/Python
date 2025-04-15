import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# tkinter로 파일 선택창 표시 (실험노트 및 조회통계 파일 선택)
Tk().withdraw()  # GUI 창을 숨김
experiment_file_path = filedialog.askopenfilename(title="실험노트 파일을 선택하세요", filetypes=[("Excel files", "*.xlsx")])
result_file_path = filedialog.askopenfilename(title="조회통계 파일을 선택하세요", filetypes=[("Excel files", "*.xlsx")])

# 데이터를 읽어오기 (사용자가 선택한 실험노트, 조회통계 자료 excel 파일 경로 이용)
data_df = pd.read_excel(experiment_file_path)
result_df = pd.read_excel(result_file_path)

# 데이터를 읽어오기(실험노트, 조회통계 자료 excel통합문서 양식으로 저장한 다음 저장)
# data_df = pd.read_excel('실험노트.xlsx')
# result_df = pd.read_excel('조회통계.xlsx')

# '접수번호'를 기준으로 데이터 병합
merged_df = pd.merge(data_df, result_df, on='접수번호', how='inner')

# 병합된 데이터를 Excel 파일로 저장
merged_df.to_excel('1.병합 데이터.xlsx', index=False)

# 병합 데이터에서 '실내공기질'이 포함된 데이터만 필터링
filtered_df = merged_df[merged_df['검체유형_x'].str.contains('실내공기질', na=False)]

# 필터링된 데이터 확인
print(filtered_df.head())

# 대기 제외, 실내공기질만 필터링된 데이터를 Excel 파일로 저장
filtered_df.to_excel('2.실내공기질 필터링 데이터.xlsx', index=False)

# '시설군' 컬럼 추가 (검체유형_x 열에서 파생)
filtered_df['시설군'] = filtered_df['검체유형_x'].str.split('/').str[2]

# 지정된 열 선택(필요한 변수만 추출)
selected_columns = [
    '의뢰기관', '시설군', '배출시설', '시설명', '채취장소_x', '시료명_x',
    'PM10', 'PM2.5', 'CO2', '폼알데하이드', '부유세균', '라돈', '라돈(밀폐)', '벤젠', '톨루엔', '에틸벤젠', '자일렌', '스틸렌', '부적합항목', '확인일', '검사결과','접수번호', '접수일자'
]

# 해당 열만 포함된 새 DataFrame 생성
extracted_df = filtered_df[selected_columns]

# 필요한 경우, 데이터를 Excel 파일로 저장
extracted_df.to_excel('3.추출 데이터.xlsx', index=False)

# 데이터 확인
print(extracted_df.head())

# 엑셀 파일 로드
file_path = '3.추출 데이터.xlsx'
df = pd.read_excel(file_path, sheet_name='Sheet1')

# 열 이름 변경
rename_columns = {
    '의뢰기관': '시군명',
    '배출시설': '세부시설군',
    '채취장소_x': '주소',
    'PM10': '미세먼지(PM10)',
    'PM2.5': '초미세먼지(PM2.5)',
    'CO2': '이산화탄소',
    '부유세균': '총부유세균'
}

# '시설군'에서 '기존공동주택'과 '신축공동주택'만 필터링
filtered_df = df[df['시설군'].isin(['기존공동주택', '신축공동주택'])]

# 열 순서 정의 및 데이터 정렬
desired_column_order = [
    '시군명', '시설군', '시설명', '세부시설군', '주소', '검사결과', '부적합항목', '접수일자',
    '확인일', '접수번호_10자리', '접수번호', '채취지점', '미세먼지(PM10)', '초미세먼지(PM2.5)',
    '이산화탄소', '폼알데하이드', '총부유세균', '라돈', '라돈(밀폐)', '벤젠', '톨루엔',
    '에틸벤젠', '자일렌', '스틸렌'
]

# 실제 데이터에 존재하는 열만 선택하도록 처리
filtered_df = filtered_df[[col for col in desired_column_order if col in filtered_df.columns]]

# 필터링된 데이터 출력 또는 저장
filtered_df.to_excel('4.기존신축공동주택 데이터.xlsx', index=False)

print(filtered_df)

# 데이터 파일 로드
file_path = '3.추출 데이터.xlsx'
data = pd.read_excel(file_path)

# 열 이름 변경
rename_columns = {
    '의뢰기관': '시군명',
    '배출시설': '세부시설군',
    '채취장소_x': '주소',
    'PM10': '미세먼지(PM10)',
    'PM2.5': '초미세먼지(PM2.5)',
    'CO2': '이산화탄소',
    '부유세균': '총부유세균'
}
data.rename(columns=rename_columns, inplace=True)

# 텍스트 데이터를 숫자로 변환 및 "불검출" 처리
columns_to_clean = ['미세먼지(PM10)', '초미세먼지(PM2.5)', '이산화탄소', '폼알데하이드', '총부유세균', '라돈', '벤젠', '톨루엔', '에틸벤젠', '자일렌', '스틸렌']
for col in columns_to_clean:
    if col in data.columns:
        data[col] = data[col].replace("불검출", None)  # "불검출"을 NaN으로 처리
        data[col] = pd.to_numeric(data[col], errors='coerce')  # 숫자로 변환 가능한 값만 남김

# 접수번호의 앞 10자리를 기준으로 새로운 열 생성
data['접수번호_10자리'] = data['접수번호'].str[:10]

# 수치 데이터의 평균 계산 (접수번호의 앞 10자리를 기준으로)
numeric_columns = ['미세먼지(PM10)', '초미세먼지(PM2.5)', '이산화탄소', '폼알데하이드', '총부유세균', '라돈']
data_avg = data.groupby(['시설명', '접수번호_10자리']).agg({col: 'mean' for col in numeric_columns if col in data.columns}).reset_index()

# 시료명 데이터 결합 및 채취지점 생성
data['채취지점'] = data.groupby('시설명')['시료명_x'].transform(lambda x: ', '.join(x.unique()))

# 채취지점 데이터와 합치기
columns_to_drop = numeric_columns + ['미세먼지(PM10_1)', '초미세먼지(PM2.5_1)', '미세먼지(PM10_2)', '초미세먼지(PM2.5_2)', 'CO2', '이산화탄소']
columns_to_drop = [col for col in columns_to_drop if col in data.columns]  # 실제 존재하는 열만 제거
data_merged = data_avg.merge(data.drop(columns=columns_to_drop).drop_duplicates(['시설명', '접수번호_10자리']), on=['시설명', '접수번호_10자리'])

# 불필요한 시료명_x 열 제거
if '시료명_x' in data_merged.columns:
    data_merged.drop(columns='시료명_x', inplace=True)

# 1. 검사결과가 '적합'이면 부적합항목을 빈칸으로 설정
data_merged.loc[data_merged['검사결과'] == '적합', '부적합항목'] = ""

# 2. '시설군'에서 '기존공동주택', '신축공동주택' 제거
data_filtered = data_merged[~data_merged['시설군'].isin(['기존공동주택', '신축공동주택'])]

# 열 순서 정의 및 데이터 정렬
desired_column_order = [
    '시군명', '시설군', '시설명', '세부시설군', '주소', '검사결과', '부적합항목', '접수일자',
    '확인일', '접수번호_10자리', '접수번호', '채취지점', '미세먼지(PM10)', '초미세먼지(PM2.5)',
    '이산화탄소', '폼알데하이드', '총부유세균', '라돈', '라돈(밀폐)', '벤젠', '톨루엔',
    '에틸벤젠', '자일렌', '스틸렌'
]

# 열 순서 변경 및 데이터 정렬
data_sorted = data_filtered[desired_column_order].sort_values(by=['시군명', '시설군', '세부시설군', '시설명', '주소', '채취지점'])

# 결과 파일로 저장 (xlsx 파일로 저장)
output_path = '5.다중이용시설 데이터.xlsx'
data_sorted.to_excel(output_path, index=False)

# 두 개의 엑셀 파일 로드
file_1 = '4.기존신축공동주택 데이터.xlsx'
file_2 = '5.다중이용시설 데이터.xlsx'

# 각 파일의 데이터를 불러옴
df1 = pd.read_excel(file_1)
df2 = pd.read_excel(file_2)

# 1. 숫자 데이터를 올바른 형식으로 변환
def convert_numeric_with_preservation(df, columns):
    for col in columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').combine_first(df[col])  # 숫자로 변환 가능한 값은 숫자로 변환, 나머지는 그대로 유지
    return df

# 변환할 열 리스트
columns_to_convert = ['벤젠', '톨루엔', '에틸벤젠', '자일렌', '스틸렌']

# 기존신축공동주택 시트 데이터에 변환 적용
df1 = convert_numeric_with_preservation(df1, columns_to_convert)

# 2. 확인일 변수를 날짜 형식으로 변환하고, 시간 제거 및 데이터프레임을 오름차순으로 정렬
def sort_by_date(df, date_column):
    # '확인일'을 날짜 형식으로 변환 (시간 제거)
    df[date_column] = pd.to_datetime(df[date_column], errors='coerce').dt.date  # 날짜만 남기고 시간 제거
    df = df.sort_values(by=date_column, ascending=True)  # 오름차순 정렬
    return df

# 확인일 기준 정렬 적용
df1 = sort_by_date(df1, '확인일')
df2 = sort_by_date(df2, '확인일')

# 3. 다중이용시설 데이터에서 "접수번호" 변수 제거
if '접수번호' in df2.columns:
    df2 = df2.drop(columns=['접수번호'])

# 결과를 하나의 엑셀 파일로 저장 (두 개의 시트에 각각 저장)
output_file = '6.통합 데이터.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df2.to_excel(writer, sheet_name='다중이용시설', index=False)  # 결과를 시트1에 저장
    df1.to_excel(writer, sheet_name='기존신축공동주택', index=False)  # 기존, 신축공동주택 데이터를 시트2에 저장

# 저장된 엑셀 파일을 다시 불러옴
wb = load_workbook(output_file)

# 각 시트에 대해 서식 지정
for sheet in wb.sheetnames:
    ws = wb[sheet]

    # 글자 크기를 10으로 설정
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(size=10)

    # 모든 열의 너비를 11로 고정
    for col in ws.columns:
        column = col[0].column_letter  # 열 이름 가져오기 (A, B, C...)
        ws.column_dimensions[column].width = 11  # 열 너비 고정

# 파일 저장
wb.save(output_file)

print("엑셀 파일이 서식과 함께 '6.통합 데이터.xlsx'로 저장되었습니다.")