# C:\Git> # 가상환경 활성화 (PowerShell용)
# cd "c:/Git/Python"
# .venv\Scripts\Activate.ps1
# cd "C:\Git\excel\2.실내공기질실적정리\3.프로그램(배포형)"
# pyinstaller --noconfirm --onefile --windowed --name="실내공기질_처리기" 0.vscode_indoor_air_app.py

# -*- coding: utf-8 -*-

import pandas as pd
from tkinter import filedialog, Tk, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import sys


def convert_numeric_with_preservation(df, columns):
    for col in columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').combine_first(df[col])
    return df


def sort_by_date(df, date_column):
    df[date_column] = pd.to_datetime(df[date_column], errors='coerce').dt.date
    df = df.sort_values(by=date_column, ascending=True)
    return df


def save_excel(df, filename):
    df.to_excel(filename, index=False)
    print(f"📁 저장됨: {filename}")


def main():
    root = Tk()
    root.withdraw()

    messagebox.showinfo("실내공기질 데이터 처리", "실험노트 파일을 선택하세요")
    note_path = filedialog.askopenfilename(title="실험노트", filetypes=[("Excel files", "*.xlsx")])
    if not note_path:
        sys.exit()

    messagebox.showinfo("실내공기질 데이터 처리", "조회통계 파일을 선택하세요")
    stat_path = filedialog.askopenfilename(title="조회통계", filetypes=[("Excel files", "*.xlsx")])
    if not stat_path:
        sys.exit()

    output_dir = "결과_파일"
    os.makedirs(output_dir, exist_ok=True)

    # 병합 및 필터링
    data_df = pd.read_excel(note_path)
    result_df = pd.read_excel(stat_path)
    merged_df = pd.merge(data_df, result_df, on='접수번호', how='inner')
    save_excel(merged_df, os.path.join(output_dir, "1.병합 데이터.xlsx"))

    filtered_df = merged_df[merged_df['검체유형_x'].str.contains('실내공기질', na=False)].copy()
    save_excel(filtered_df, os.path.join(output_dir, "2.실내공기질 필터링 데이터.xlsx"))

    filtered_df['시설군'] = filtered_df['검체유형_x'].str.split('/').str[2]
    selected_columns = [
        '의뢰기관', '시설군', '배출시설', '시설명', '채취장소_x', '시료명_x',
        'PM10', 'PM2.5', 'CO2', '폼알데하이드', '부유세균', '일산화탄소', '라돈', '라돈(밀폐)',
        '벤젠', '톨루엔', '에틸벤젠', '자일렌', '스틸렌', '부적합항목', '확인일', '검사결과', '접수번호', '접수일자'
    ]
    extracted_df = filtered_df[selected_columns].copy()
    save_excel(extracted_df, os.path.join(output_dir, "3.추출 데이터.xlsx"))

    # 공통 전처리
    df = extracted_df.rename(columns={
        '의뢰기관': '시군명', '배출시설': '세부시설군', '채취장소_x': '주소',
        'PM10': '미세먼지(PM10)', 'PM2.5': '초미세먼지(PM2.5)', 'CO2': '이산화탄소',
        '부유세균': '총부유세균', '시료명_x': '시료명'
    })
    df['접수번호_10자리'] = df['접수번호'].astype(str).str[:10]
    df['채취지점'] = df.groupby('시설명')['시료명'].transform(lambda x: ', '.join(x.unique()))

    # 기존/신축공동주택
    apart_df = df[df['시설군'].isin(['기존공동주택', '신축공동주택'])].copy()
    save_excel(apart_df, os.path.join(output_dir, "4.기존신축공동주택 데이터.xlsx"))

    # 다중이용시설
    for col in ['미세먼지(PM10)', '초미세먼지(PM2.5)', '이산화탄소', '폼알데하이드', '총부유세균', '일산화탄소', '라돈']:
        if col in df.columns:
            df[col] = df[col].replace("불검출", None)
            df[col] = pd.to_numeric(df[col], errors='coerce')

    avg_cols = ['미세먼지(PM10)', '초미세먼지(PM2.5)', '이산화탄소', '폼알데하이드', '총부유세균', '일산화탄소', '라돈']
    avg_df = df.groupby(['시설명', '접수번호_10자리'])[avg_cols].mean().reset_index()
    reduced = df.drop(columns=avg_cols, errors='ignore').drop_duplicates(['시설명', '접수번호_10자리'])
    merged = avg_df.merge(reduced, on=['시설명', '접수번호_10자리'])
    merged.loc[merged['검사결과'] == '적합', '부적합항목'] = ""
    multi_df = merged[~merged['시설군'].isin(['기존공동주택', '신축공동주택'])].copy()
    save_excel(multi_df, os.path.join(output_dir, "5.다중이용시설 데이터.xlsx"))

    # 통합
    apart_df = pd.read_excel(os.path.join(output_dir, "4.기존신축공동주택 데이터.xlsx"))
    multi_df = pd.read_excel(os.path.join(output_dir, "5.다중이용시설 데이터.xlsx"))

    columns_to_convert = ['벤젠', '톨루엔', '에틸벤젠', '자일렌', '스틸렌']
    apart_df = convert_numeric_with_preservation(apart_df, columns_to_convert)

    apart_df = sort_by_date(apart_df, '확인일')
    multi_df = sort_by_date(multi_df, '확인일')
    if '접수번호' in multi_df.columns:
        multi_df = multi_df.drop(columns='접수번호')

    output_file = os.path.join(output_dir, '6.통합 데이터.xlsx')
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        multi_df.to_excel(writer, sheet_name='다중이용시설', index=False)
        apart_df.to_excel(writer, sheet_name='기존신축공동주택', index=False)

    wb = load_workbook(output_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(size=10)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 11
    wb.save(output_file)

    done_window = Tk()
    done_window.title("작업 완료")
    done_window.geometry("300x150")

    from tkinter import Label, Button

    Label(done_window, text="모든 단계가 완료되었습니다!\n\n결과는 '결과_파일' 폴더에 저장되었습니다.", padx=20, pady=20).pack()

    def terminate():
        done_window.destroy()
        root.quit()
        root.destroy()
        os._exit(0)  # 완전 종료

    Button(done_window, text="종료", command=terminate, padx=10, pady=5).pack(pady=10)

    done_window.mainloop()

if __name__ == "__main__":
    main()
