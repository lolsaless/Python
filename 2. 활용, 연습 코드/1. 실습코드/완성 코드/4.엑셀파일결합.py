import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook


def merge_excel_files(folder_path):
    # 새로운 엑셀 파일 생성
    merged_workbook = Workbook()
    merged_workbook.remove(merged_workbook.active)  # 기본 시트 제거

    # 폴더 내의 모든 파일 검색
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            sheet_name = os.path.splitext(filename)[0]

            # 엑셀 파일 열기
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            # 새로운 시트로 데이터 복사
            new_sheet = merged_workbook.create_sheet(title=sheet_name)
            for row in worksheet.iter_rows(values_only=True):
                new_sheet.append(row)

            workbook.close()

    # 결과 저장
    result_file_path = os.path.join(folder_path, 'merged.xlsx')
    merged_workbook.save(result_file_path)
    merged_workbook.close()

    print(f"파일이 성공적으로 병합되었습니다. 저장 위치: {result_file_path}")


# 폴더 선택 대화상자 열기
root = tk.Tk()
root.withdraw()
folder_path = filedialog.askdirectory(title="폴더를 선택하세요.")

# 폴더 선택이 이루어진 경우에만 실행
if folder_path:
    merge_excel_files(folder_path)
else:
    print("폴더 선택이 취소되었습니다.")
