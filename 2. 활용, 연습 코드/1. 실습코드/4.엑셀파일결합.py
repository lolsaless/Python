import os
import glob
from openpyxl import Workbook, load_workbook
from tkinter import Tk
from tkinter.filedialog import askdirectory

# 폴더 선택 대화상자를 통해 폴더 선택
Tk().withdraw()
folder_path = askdirectory(title='폴더 선택')

# 새로운 엑셀 파일 생성
merged_workbook = Workbook()

# 폴더 내의 엑셀 파일 목록 가져오기
excel_files = glob.glob(os.path.join(folder_path, '*.xlsx'))

for file in excel_files:
    # 엑셀 파일 로드
    workbook = load_workbook(file)
    
    # 파일 이름에서 확장자 제거하여 시트명으로 사용
    sheet_name = os.path.splitext(os.path.basename(file))[0]
    
    # 새로운 시트 생성
    merged_workbook.create_sheet(title=sheet_name)
    
    # 원본 엑셀 파일의 시트 데이터를 새로운 시트로 복사
    for sheet in workbook.sheetnames:
        source_sheet = workbook[sheet]
        target_sheet = merged_workbook[sheet_name]
        
        for row in source_sheet.iter_rows(values_only=True):
            target_sheet.append(row)
    
    # 원본 엑셀 파일 닫기
    workbook.close()

# 기본 시트 삭제
merged_workbook.remove(merged_workbook.active)

# 결과를 저장할 파일명 입력
output_filename = input("저장할 파일명을 입력하세요: ") + ".xlsx"

# 결과 파일 저장
merged_workbook.save(output_filename)
merged_workbook.close()

print("병합이 완료되었습니다.")
